import openpyxl


class Page:
    datavalue = [{"uname": "Lijo.mathew+129@careerswales.gov.wales", "passwd": "Lijo@cardiff2"}]


"""
  @staticmethod
    def gettestdata(testcase_name):

        Dict = {}

        book = openpyxl.load_workbook("C:\\Users\\Lijo.mathew\\PycharmProjects\\Pytest_demo\\Pytest.xlsx")

        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
"""



"""
    @staticmethod
    def gettestdata_full():

        Dict1 = {}
        Dict2 = {}

        book = openpyxl.load_workbook("C:\\Users\\Lijo.mathew\\PycharmProjects\\Pytest_demo\\Pytest.xlsx")

        sheet = book.active

        for i in range(2, sheet.max_row + 1):
            for j in range(2, sheet.max_column + 1):
                if i == 2:
                    Dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                elif i == 3:
                    Dict2[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict1,Dict2]
"""

